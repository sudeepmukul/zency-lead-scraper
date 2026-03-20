#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════╗
║        LeadGen Pro — Intelligent Business Lead Scraper           ║
║        Built for Digital Marketing Agencies                      ║
║                                                                  ║
║  Features:                                                       ║
║  • Google Maps scraping (zero API cost)                          ║
║  • Auto email, phone, social media extraction                    ║
║  • AI lead scoring (0–100) with grade & signals                  ║
║  • Business size detection from review count                     ║
║  • Filtered, ranked output (Excel + CSV)                         ║
╚══════════════════════════════════════════════════════════════════╝

SETUP (run once):
    pip install playwright playwright-stealth requests beautifulsoup4 openpyxl rich
    playwright install chromium

USAGE:
    python leadgen_pro.py
"""

from __future__ import annotations

import csv
import os
import re
import sys
import time
import logging
from pathlib import Path
from urllib.parse import urlparse, urljoin, quote_plus

import requests
from bs4 import BeautifulSoup

# ── Optional imports ───────────────────────────────────────────────────────

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

try:
    from playwright_stealth import stealth_sync
    STEALTH_AVAILABLE = True
except ImportError:
    STEALTH_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn, TimeElapsedColumn
    from rich.table import Table
    from rich.text import Text
    RICH_AVAILABLE = True
    console = Console()
except ImportError:
    RICH_AVAILABLE = False
    console = None

logging.basicConfig(level=logging.WARNING)
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')

SOCIAL_RE = {
    'instagram': re.compile(
        r'(?:https?://)?(?:www\.)?instagram\.com/(?!p/|reel/|explore/|stories/)([A-Za-z0-9._]{1,30})/?', re.I),
    'facebook': re.compile(
        r'(?:https?://)?(?:www\.)?facebook\.com/(?!sharer|share|dialog|login|photo|video|watch|groups|events|pages/create)([A-Za-z0-9._\-]{3,80})/?', re.I),
    'twitter': re.compile(
        r'(?:https?://)?(?:www\.)?(?:twitter|x)\.com/(?!intent/|share|home|search)([A-Za-z0-9_]{1,15})/?', re.I),
    'linkedin': re.compile(
        r'(?:https?://)?(?:www\.)?linkedin\.com/(company|in)/([A-Za-z0-9._\-]+)/?', re.I),
    'youtube': re.compile(
        r'(?:https?://)?(?:www\.)?youtube\.com/(?:channel/|c/|user/|@)([A-Za-z0-9._\-]+)/?', re.I),
}

WEBSITE_BUILDERS = {
    'wix.com', 'squarespace.com', 'weebly.com', 'wordpress.com',
    'webflow.io', 'godaddy.com', 'jimdo.com', 'strikingly.com',
    'site123.com', 'webnode.com', 'yola.com',
}

SOCIAL_DOMAINS = {
    'instagram.com', 'facebook.com', 'twitter.com', 'x.com',
    'linkedin.com', 'youtube.com', 'tiktok.com', 'pinterest.com',
}

SKIP_DOMAINS = {
    'google.com', 'maps.google.com', 'goo.gl', 'googleapis.com',
} | SOCIAL_DOMAINS

BAD_EMAIL_DOMAINS = {
    'example.com', 'domain.com', 'email.com', 'test.com',
    'sentry.io', 'wixpress.com', 'squarespace.com',
}

BAD_EMAIL_KEYWORDS = ['.png', '.jpg', '.gif', '.svg', '.css', '.js',
                      'noreply', 'no-reply', 'donotreply', 'bounce', 'mailer']

HEADERS = {
    'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                   'AppleWebKit/537.36 (KHTML, like Gecko) '
                   'Chrome/122.0.0.0 Safari/537.36'),
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}

# NIL sentinel used throughout
NIL = 'nil'


# ══════════════════════════════════════════════════════════════════════════════
# LEAD SCORING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def score_lead(data: dict) -> dict:
    """
    Score a lead 0–100 based on how likely they are to need & pay for
    digital marketing services (web design, social media, branding, ads).

    Logic: gaps in digital presence = more opportunity for your agency.
    """
    score = 0
    signals = []   # Short opportunity tags shown in output
    reasons = []   # Detailed breakdown

    website = (data.get('website') or '').strip()
    has_real_website = bool(website) and not any(d in website for d in SOCIAL_DOMAINS)

    # ── 1. WEBSITE (0–25 pts) ─────────────────────────────────────────────
    if not has_real_website:
        score += 25
        signals.append('No Website')
        reasons.append('🔴 No website → prime web design pitch')
    else:
        builder_used = next((b for b in WEBSITE_BUILDERS if b in website), None)
        if builder_used:
            score += 15
            signals.append(f'DIY Site ({builder_used.split(".")[0].title()})')
            reasons.append(f'🟡 Uses {builder_used} → upgrade opportunity')
        else:
            score += 5
            reasons.append('🟢 Has custom website')

    # ── 2. SOCIAL MEDIA (0–20 pts) ────────────────────────────────────────
    instagram  = data.get('instagram', NIL) or NIL
    facebook   = data.get('facebook',  NIL) or NIL
    twitter    = data.get('twitter',   NIL) or NIL
    linkedin   = data.get('linkedin',  NIL) or NIL

    present = sum(1 for s in [instagram, facebook, twitter, linkedin] if s != NIL)

    if present == 0:
        score += 20
        signals.append('No Social Media')
        reasons.append('🔴 Zero social presence → huge content opportunity')
    elif present == 1:
        score += 13
        signals.append('Weak Social Presence')
        reasons.append('🟡 Only 1 social channel → needs more platforms')
    elif present == 2:
        score += 7
        reasons.append('🟡 2 social channels → room to grow')
    else:
        score += 2
        reasons.append('🟢 Active social presence')

    # ── 3. EMAIL TYPE (0–12 pts) ──────────────────────────────────────────
    email = (data.get('email') or '').lower().strip()
    personal_providers = {'gmail.com', 'yahoo.com', 'hotmail.com',
                          'outlook.com', 'rediffmail.com', 'ymail.com'}
    if not email or email == NIL:
        score += 4
        signals.append('No Email Found')
        reasons.append('⚪ No email found (harder to reach)')
    else:
        domain = email.split('@')[-1]
        if domain in personal_providers:
            score += 10
            signals.append('Personal Email (Gmail/Yahoo)')
            reasons.append('🟡 Personal email → no professional branding')
        else:
            score += 4
            reasons.append('🟢 Professional email domain')

    # ── 4. BUSINESS SIZE from review count (0–15 pts) ────────────────────
    raw_reviews = data.get('reviews_count', 0) or 0
    try:
        review_count = int(str(raw_reviews).replace(',', '').replace('+', '').strip() or 0)
    except (ValueError, TypeError):
        review_count = 0

    if review_count == 0:
        score += 5
        data['business_size'] = 'Unknown'
        reasons.append('⚪ No reviews yet (brand new or inactive GMB)')
    elif review_count < 15:
        score += 15
        data['business_size'] = 'Micro'
        signals.append('Micro Business')
        reasons.append('🔵 Micro business (<15 reviews) → very approachable')
    elif review_count < 75:
        score += 12
        data['business_size'] = 'Small'
        reasons.append('🔵 Small business → solid agency fit')
    elif review_count < 300:
        score += 8
        data['business_size'] = 'Medium'
        reasons.append('🟡 Medium business → may have some marketing in place')
    elif review_count < 1000:
        score += 4
        data['business_size'] = 'Large'
        reasons.append('🟠 Large business → may have in-house team')
    else:
        score += 1
        data['business_size'] = 'Enterprise'
        signals.append('Enterprise (Hard Sell)')
        reasons.append('🔴 Enterprise business → low agency conversion rate')

    # ── 5. REACHABILITY (0–12 pts) ───────────────────────────────────────
    has_phone = (data.get('phone') or NIL) != NIL
    has_email = email not in ('', NIL)

    if has_phone and has_email:
        score += 12
        reasons.append('✅ Fully reachable (phone + email)')
    elif has_phone:
        score += 7
        reasons.append('📞 Phone only — no email found')
    elif has_email:
        score += 8
        reasons.append('📧 Email only — no phone found')
    else:
        score -= 8
        signals.append('Hard to Reach')
        reasons.append('❌ No contact info found')

    # ── 6. RATING / REPUTATION (0–8 pts) ─────────────────────────────────
    try:
        rating = float(str(data.get('rating') or 0).replace(',', '.'))
    except (ValueError, TypeError):
        rating = 0.0

    if 3.0 <= rating < 4.0:
        score += 6
        signals.append('Low Rating (Reputation Help)')
        reasons.append('⚠️ Below-average rating → reputation mgmt pitch')
    elif 4.0 <= rating < 4.7:
        score += 8
        reasons.append('⭐ Good rating → active, stable business')
    elif rating >= 4.7:
        score += 4
        reasons.append('🌟 Excellent rating → established business')
    elif rating > 0:
        score += 2
        reasons.append(f'Rating: {rating}')

    # ── CLAMP & GRADE ─────────────────────────────────────────────────────
    score = max(0, min(score, 100))

    if score >= 72:
        grade = '🔥 HOT'
    elif score >= 52:
        grade = '⚡ WARM'
    elif score >= 32:
        grade = '🌡️ COOL'
    else:
        grade = '❄️ COLD'

    return {
        'lead_score':         score,
        'lead_grade':         grade,
        'opportunity_signals': ', '.join(signals) if signals else 'Well Established',
        'score_breakdown':    ' | '.join(reasons),
    }


# ══════════════════════════════════════════════════════════════════════════════
# CONTACT EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def _clean_email(raw: str) -> str | None:
    email = raw.lower().strip().strip('.')
    if len(email) < 6 or len(email) > 100:
        return None
    domain = email.split('@')[-1]
    if domain in BAD_EMAIL_DOMAINS:
        return None
    if any(kw in email for kw in BAD_EMAIL_KEYWORDS):
        return None
    if email.count('@') != 1:
        return None
    return email


def _clean_social_handle(platform: str, match) -> str | None:
    """Validate and format a social media match."""
    if platform == 'linkedin':
        kind, handle = match.group(1), match.group(2)
        base = 'linkedin.com'
        return f'https://{base}/{kind}/{handle.rstrip("/")}'
    else:
        handle = match.group(1).rstrip('/')
        blocked = {
            'share', 'sharer', 'intent', 'dialog', 'login', 'signup',
            'about', 'help', 'legal', 'policies', 'privacy', 'terms',
            'explore', 'feed', 'home', 'search', 'notifications',
        }
        if handle.lower() in blocked or len(handle) < 2:
            return None
        domain_map = {
            'twitter':   'x.com',
            'instagram': 'instagram.com',
            'facebook':  'facebook.com',
            'youtube':   'youtube.com',
        }
        base = domain_map.get(platform, f'{platform}.com')
        return f'https://{base}/{handle}'


def extract_contacts(html: str, base_url: str = '') -> dict:
    """Parse HTML and extract every useful contact signal."""
    soup = BeautifulSoup(html, 'html.parser')

    emails, phones = [], []
    socials = {k: NIL for k in SOCIAL_RE}

    # ── <a href> links ─────────────────────────────────────────────────────
    for a in soup.find_all('a', href=True):
        href = (a.get('href') or '').strip()

        if href.startswith('mailto:'):
            raw = href[7:].split('?')[0].strip()
            if '@' in raw:
                cleaned = _clean_email(raw)
                if cleaned:
                    emails.append(cleaned)

        elif href.startswith('tel:'):
            phone = re.sub(r'[^\d+\-\s()]', '', href[4:]).strip()
            if len(phone) >= 7:
                phones.append(phone)

        else:
            full_href = urljoin(base_url, href) if not href.startswith('http') else href
            for platform, pattern in SOCIAL_RE.items():
                if socials[platform] == NIL:
                    m = pattern.search(full_href)
                    if m:
                        clean = _clean_social_handle(platform, m)
                        if clean:
                            socials[platform] = clean

    # ── Full text scan for emails ──────────────────────────────────────────
    for tag in soup(['script', 'style', 'noscript', 'meta']):
        tag.decompose()
    text = soup.get_text(separator=' ')

    for raw in EMAIL_RE.findall(text):
        cleaned = _clean_email(raw)
        if cleaned:
            emails.append(cleaned)

    # ── Deduplicate ────────────────────────────────────────────────────────
    emails = list(dict.fromkeys(emails))
    phones = list(dict.fromkeys(phones))

    return {'emails': emails, 'phones': phones, **socials}


def scrape_website_contacts(url: str, pw=None) -> dict:
    """
    Fetch a business website (and its /contact page) to extract:
    email, all_emails, social handles, website_phone.
    Falls back to Playwright for JS-heavy sites.
    """
    empty = {k: NIL for k in ['email', 'all_emails', 'instagram', 'facebook',
                                'twitter', 'linkedin', 'youtube', 'website_phone']}
    if not url:
        return empty

    all_emails, all_phones = [], []
    socials = {k: NIL for k in SOCIAL_RE}

    pages = [url, urljoin(url, '/contact'), urljoin(url, '/contact-us'),
             urljoin(url, '/about'), urljoin(url, '/about-us')]

    for page_url in pages[:3]:
        try:
            r = requests.get(page_url, headers=HEADERS,
                             timeout=12, allow_redirects=True)
            if r.status_code == 200:
                contacts = extract_contacts(r.text, page_url)
                all_emails.extend(contacts['emails'])
                all_phones.extend(contacts['phones'])
                for p in SOCIAL_RE:
                    if socials[p] == NIL and contacts[p] != NIL:
                        socials[p] = contacts[p]
        except Exception:
            pass

        # Stop early if we already have good data
        if all_emails and sum(1 for v in socials.values() if v != NIL) >= 2:
            break

    # Playwright fallback if nothing found on homepage
    if not all_emails and pw:
        try:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page(user_agent=HEADERS['User-Agent'])
            if STEALTH_AVAILABLE:
                stealth_sync(page)
            page.goto(url, timeout=25_000, wait_until='domcontentloaded')
            page.wait_for_timeout(2_500)
            html = page.content()
            browser.close()

            contacts = extract_contacts(html, url)
            all_emails.extend(contacts['emails'])
            all_phones.extend(contacts['phones'])
            for p in SOCIAL_RE:
                if socials[p] == NIL and contacts[p] != NIL:
                    socials[p] = contacts[p]
        except Exception:
            pass

    all_emails = list(dict.fromkeys(all_emails))
    all_phones = list(dict.fromkeys(all_phones))

    return {
        'email':         all_emails[0] if all_emails else NIL,
        'all_emails':    '; '.join(all_emails) if all_emails else NIL,
        'website_phone': all_phones[0] if all_phones else NIL,
        **socials,
    }


# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE MAPS SCRAPER
# ══════════════════════════════════════════════════════════════════════════════

class GoogleMapsScraper:
    """Scrapes Google Maps search results without any API keys."""

    FEED_SELECTOR   = 'div[role="feed"]'
    # Multiple fallback selectors for business cards
    CARD_SELECTORS  = [
        'div[role="feed"] > div > div[jsaction*="mouseover"]',
        'div[role="feed"] a[href*="/maps/place/"]',
    ]

    def __init__(self, pw, headless: bool = True):
        self.pw       = pw
        self.headless = headless
        self.browser  = None
        self.page     = None

    def _launch(self):
        self.browser = self.pw.chromium.launch(headless=self.headless)
        ctx = self.browser.new_context(
            user_agent=HEADERS['User-Agent'],
            viewport={'width': 1366, 'height': 768},
            locale='en-US',
            timezone_id='Asia/Kolkata',
        )
        self.page = ctx.new_page()
        if STEALTH_AVAILABLE:
            stealth_sync(self.page)

    def _close(self):
        try:
            if self.browser:
                self.browser.close()
        except Exception:
            pass

    # ── Public API ────────────────────────────────────────────────────────

    def search(self, query: str, max_results: int = 20) -> list:
        """Search Google Maps and return a list of business data dicts."""
        self._launch()
        results = []
        try:
            results = self._run_search(query, max_results)
        except Exception as e:
            log.error(f'Maps search crashed: {e}')
        finally:
            self._close()
        return results

    # ── Internal ──────────────────────────────────────────────────────────

    def _run_search(self, query: str, max_results: int) -> list:
        url = f'https://www.google.com/maps/search/{quote_plus(query)}'
        self.page.goto(url, timeout=35_000, wait_until='domcontentloaded')
        self.page.wait_for_timeout(3_000)

        # Dismiss consent dialog if present
        for btn_text in ['Accept all', 'Reject all', 'Accept']:
            try:
                self.page.click(f'button:has-text("{btn_text}")', timeout=2_500)
                break
            except Exception:
                pass

        try:
            self.page.wait_for_selector(self.FEED_SELECTOR, timeout=12_000)
        except Exception:
            print('  ⚠️  Could not find Maps results feed. '
                  'Google may be blocking. Try running with headless=False.')
            return []

        results        = []
        seen_names     = set()
        scroll_count   = 0
        max_scrolls    = max(40, max_results * 2)   # generous scrolling budget
        no_new_streak  = 0

        while len(results) < max_results and scroll_count < max_scrolls:
            prev_count = len(results)

            # ── Collect visible cards ──────────────────────────────────────
            cards = self._get_cards()

            for card in cards:
                if len(results) >= max_results:
                    break
                try:
                    # Get name BEFORE clicking
                    name = self._card_name(card)
                    if not name or name in seen_names:
                        continue
                    seen_names.add(name)

                    # Click card → side panel opens
                    card.scroll_into_view_if_needed()
                    card.click()
                    self.page.wait_for_timeout(1_800)

                    biz = self._extract_panel(name)
                    if biz:
                        results.append(biz)
                        self._print_progress(len(results), max_results, biz)

                except Exception as e:
                    log.debug(f'Card error: {e}')
                    continue

            # ── Scroll feed ───────────────────────────────────────────────
            feed = self.page.query_selector(self.FEED_SELECTOR)
            if feed:
                feed.evaluate('el => el.scrollBy(0, 900)')
                self.page.wait_for_timeout(2_000)

            # Check for end-of-list marker
            try:
                end_el = self.page.query_selector('p.fontBodyMedium')
                if end_el and 'end of the list' in (end_el.inner_text() or '').lower():
                    print('  ℹ️  Reached end of Google Maps results.')
                    break
            except Exception:
                pass

            scroll_count += 1
            if len(results) == prev_count:
                no_new_streak += 1
            else:
                no_new_streak = 0

            if no_new_streak >= 6:
                print('  ℹ️  No new results in 6 scrolls — stopping.')
                break

        return results

    def _get_cards(self) -> list:
        for sel in self.CARD_SELECTORS:
            cards = self.page.query_selector_all(sel)
            if cards:
                return cards
        return []

    def _card_name(self, card) -> str:
        for sel in ['div.qBF1Pd', 'span.fontHeadlineSmall',
                    'div.NrDZNb > div', 'span[jstcache]']:
            try:
                el = card.query_selector(sel)
                if el:
                    t = el.inner_text().strip()
                    if t:
                        return t
            except Exception:
                pass
        return ''

    def _extract_panel(self, name: str) -> dict | None:
        """Extract business details from the right-side panel."""
        try:
            self.page.wait_for_timeout(600)

            d = {
                'title':          name,
                'category':       NIL,
                'rating':         NIL,
                'reviews_count':  NIL,
                'address':        NIL,
                'phone':          NIL,
                'website':        NIL,
                'maps_url':       self.page.url,
                'hours':          NIL,
                # Filled later by website scraper
                'email':          NIL,
                'all_emails':     NIL,
                'instagram':      NIL,
                'facebook':       NIL,
                'twitter':        NIL,
                'linkedin':       NIL,
                'youtube':        NIL,
                'website_phone':  NIL,
                'business_size':  'Unknown',
                'lead_score':     0,
                'lead_grade':     NIL,
                'opportunity_signals': NIL,
                'score_breakdown':     NIL,
            }

            # Category
            for sel in ['button[jsaction*="pane.rating.category"]',
                        'span.DkEaL', 'button.DkEaL']:
                try:
                    el = self.page.query_selector(sel)
                    if el:
                        d['category'] = el.inner_text().strip()
                        break
                except Exception:
                    pass

            # Rating
            try:
                el = self.page.query_selector('div.F7nice > span[aria-hidden="true"]')
                if el:
                    d['rating'] = el.inner_text().strip()
            except Exception:
                pass

            # Reviews count
            try:
                el = self.page.query_selector('div.F7nice span[aria-label*="review"]')
                if el:
                    label = el.get_attribute('aria-label') or ''
                    m = re.search(r'([\d,]+)', label)
                    if m:
                        d['reviews_count'] = m.group(1).replace(',', '')
            except Exception:
                pass

            # Address
            try:
                el = self.page.query_selector('button[data-item-id="address"]')
                if el:
                    d['address'] = el.inner_text().strip()
            except Exception:
                pass

            # Phone (from Maps panel)
            try:
                el = self.page.query_selector('button[data-item-id^="phone:tel:"]')
                if el:
                    d['phone'] = el.inner_text().strip()
                else:
                    # Fallback: find any tel: link
                    el = self.page.query_selector('a[href^="tel:"]')
                    if el:
                        d['phone'] = (el.get_attribute('href') or '').replace('tel:', '').strip()
            except Exception:
                pass

            # Website
            try:
                for sel in ['a[data-item-id="authority"]',
                             'a[aria-label*="Website"]',
                             'a[aria-label*="website"]']:
                    el = self.page.query_selector(sel)
                    if el:
                        href = el.get_attribute('href') or ''
                        if href.startswith('http'):
                            d['website'] = href
                            break
            except Exception:
                pass

            # Opening hours
            try:
                el = self.page.query_selector('div[data-hide-tooltip-on-mouse-out] button')
                if el:
                    d['hours'] = el.inner_text().strip().split('\n')[0]
            except Exception:
                pass

            return d

        except Exception as e:
            log.warning(f'Panel extraction failed for {name}: {e}')
            return None

    @staticmethod
    def _print_progress(current: int, total: int, biz: dict):
        name = (biz.get('title') or 'Unknown')[:38]
        phone = biz.get('phone') or ''
        web   = biz.get('website') or ''
        tags  = []
        if phone:
            tags.append('📞')
        if web:
            tags.append('🌐')
        tag_str = ' '.join(tags) or '—'
        print(f'  [{current:>3}/{total}] {name:<38} {tag_str}')


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

COLUMNS = [
    # (header label,           dict key,              col width)
    ('Business Name',          'title',               28),
    ('Category',               'category',            18),
    ('Rating ⭐',              'rating',               9),
    ('Reviews #',              'reviews_count',       10),
    ('Business Size',          'business_size',       13),
    ('Address',                'address',             35),
    ('Phone (Google Maps)',    'phone',               18),
    ('Website',                'website',             38),
    ('Primary Email',          'email',               32),
    ('All Emails',             'all_emails',          42),
    ('Instagram',              'instagram',           38),
    ('Facebook',               'facebook',            38),
    ('Twitter / X',            'twitter',             35),
    ('LinkedIn',               'linkedin',            38),
    ('YouTube',                'youtube',             38),
    ('Website Phone',          'website_phone',       18),
    ('Lead Score',             'lead_score',          11),
    ('Lead Grade',             'lead_grade',          12),
    ('Opportunity Signals',    'opportunity_signals', 40),
    ('Maps URL',               'maps_url',            55),
    ('Score Breakdown',        'score_breakdown',     70),
    ('Hours',                  'hours',               30),
]


def _grade_fill(grade: str) -> PatternFill:
    colors = {
        'HOT':  'FF4757',
        'WARM': 'FFA502',
        'COOL': '2980b9',
        'COLD': '95a5a6',
    }
    for key, color in colors.items():
        if key in (grade or '').upper():
            return PatternFill('solid', fgColor=color)
    return PatternFill('solid', fgColor='bdc3c7')


def save_excel(rows: list, path: str):
    if not EXCEL_AVAILABLE or not rows:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leads'

    # ── Header row ─────────────────────────────────────────────────────────
    header_fill  = PatternFill('solid', fgColor='1a1a2e')
    header_font  = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (label, _, width) in enumerate(COLUMNS, 1):
        cell         = ws.cell(row=1, column=col_idx, value=label)
        cell.fill    = header_fill
        cell.font    = header_font
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 34
    ws.freeze_panes  = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}1'

    # ── Data rows ──────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 2):
        grade = str(row.get('lead_grade') or '')
        fill  = _grade_fill(grade)
        score = row.get('lead_score', 0)

        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            val  = row.get(key, NIL)
            val  = NIL if val in (None, '', 'None') else val
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            cell.alignment = Alignment(vertical='center', wrap_text=False)

            # Colour lead score & grade columns
            if key == 'lead_score':
                cell.fill      = fill
                cell.font      = Font(bold=True, size=12, color='FFFFFF', name='Calibri')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif key == 'lead_grade':
                cell.fill      = fill
                cell.font      = Font(bold=True, color='FFFFFF', name='Calibri')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # Hyperlinks
            elif key in ('website', 'instagram', 'facebook', 'twitter',
                         'linkedin', 'youtube', 'maps_url'):
                if val and val != NIL and val.startswith('http'):
                    cell.hyperlink = val
                    cell.font      = Font(color='1e90ff', underline='single')
            elif key == 'email' and val != NIL and '@' in val:
                cell.hyperlink = f'mailto:{val}'
                cell.font      = Font(color='1e90ff', underline='single')

        ws.row_dimensions[row_idx].height = 18

    wb.save(path)
    print(f'  ✅ Excel: {path}')


def save_csv(rows: list, path: str):
    if not rows:
        return
    keys = [k for _, k, _ in COLUMNS]
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=keys, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)
    print(f'  ✅ CSV:   {path}')


# ══════════════════════════════════════════════════════════════════════════════
# CLI HELPERS
# ══════════════════════════════════════════════════════════════════════════════

BANNER = r"""
 _                     _  ____
| |    ___  __ _  __| |/ ___| ___ _ __     _ __  _ __ ___
| |   / _ \/ _` |/ _` | |  _ / _ \ '_ \   | '_ \| '__/ _ \
| |__|  __/ (_| | (_| | |_| |  __/ | | |  | |_) | | | (_) |
|_____\___|\__,_|\__,_|\____|\___|_| |_|  | .__/|_|  \___/
                                           |_|
           Digital Marketing Lead Engine  v2.0
"""


def cprint(text: str, color_code: str = ''):
    """Print with ANSI color codes (works on all modern terminals)."""
    reset = '\033[0m'
    print(f'{color_code}{text}{reset}')


def ask(prompt: str, default: str = '') -> str:
    try:
        val = input(f'\033[1;36m{prompt}\033[0m').strip()
        return val if val else default
    except (KeyboardInterrupt, EOFError):
        print('\n👋 Bye!')
        sys.exit(0)


def print_summary(leads: list):
    hot  = sum(1 for l in leads if 'HOT'  in (l.get('lead_grade') or ''))
    warm = sum(1 for l in leads if 'WARM' in (l.get('lead_grade') or ''))
    cool = sum(1 for l in leads if 'COOL' in (l.get('lead_grade') or ''))
    cold = sum(1 for l in leads if 'COLD' in (l.get('lead_grade') or ''))

    with_email = sum(1 for l in leads if (l.get('email') or NIL) != NIL)
    with_phone = sum(1 for l in leads if (l.get('phone') or NIL) != NIL)
    with_ig    = sum(1 for l in leads if (l.get('instagram') or NIL) != NIL)
    with_web   = sum(1 for l in leads if (l.get('website') or NIL) != NIL)

    print('\n' + '═' * 62)
    cprint('  🎯  LEAD GENERATION COMPLETE', '\033[1;32m')
    print('═' * 62)
    print(f'  Total Leads Generated  : {len(leads)}')
    print(f'  🔥 HOT  (score ≥ 72)   : {hot}')
    print(f'  ⚡ WARM (score ≥ 52)   : {warm}')
    print(f'  🌡️  COOL (score ≥ 32)   : {cool}')
    print(f'  ❄️  COLD (score < 32)   : {cold}')
    print('  ──────────────────────────────────────────────────')
    print(f'  📧 With Email          : {with_email}')
    print(f'  📞 With Phone          : {with_phone}')
    print(f'  🌐 With Website        : {with_web}')
    print(f'  📸 With Instagram      : {with_ig}')
    print('═' * 62)

    hot_leads = [l for l in leads if 'HOT' in (l.get('lead_grade') or '')][:5]
    if hot_leads:
        print('\n  🔥 TOP HOT LEADS TO CONTACT FIRST:')
        for i, l in enumerate(hot_leads, 1):
            name    = (l.get('title') or 'N/A')[:35]
            score   = l.get('lead_score', 0)
            signals = (l.get('opportunity_signals') or '')[:50]
            email   = l.get('email') or NIL
            phone   = l.get('phone') or NIL
            print(f'  {i}. [{score:>3}/100] {name}')
            print(f'       Signals : {signals}')
            print(f'       Contact : {email}  |  {phone}')
    print()


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    cprint(BANNER, '\033[1;35m')

    # ── Dependency check ──────────────────────────────────────────────────
    if not PLAYWRIGHT_AVAILABLE:
        print('❌ Playwright not installed!')
        print('   pip install playwright && playwright install chromium')
        sys.exit(1)

    if not STEALTH_AVAILABLE:
        cprint('⚠️  playwright-stealth missing — Google may block. '
               'Run: pip install playwright-stealth\n', '\033[1;33m')

    if not EXCEL_AVAILABLE:
        cprint('⚠️  openpyxl missing — Excel output disabled. '
               'Run: pip install openpyxl\n', '\033[1;33m')

    # ── User inputs ───────────────────────────────────────────────────────
    print('─' * 62)
    query      = ask('🔍 Search query  (e.g. Cafes in Hyderabad): ')
    if not query:
        print('No query entered. Exiting.')
        sys.exit(0)

    raw_count  = ask('📊 How many leads do you want? (default 50): ', '50')
    try:
        max_leads  = max(1, int(raw_count))
    except ValueError:
        max_leads  = 50

    scrape_web = ask('🌐 Scrape websites for email/social media? [y/n] (default y): ', 'y').lower()
    do_web     = scrape_web != 'n'

    fmt        = ask('💾 Output format [excel / csv / both] (default both): ', 'both').lower()

    headless   = ask('👁️  Run browser in background? [y/n] (default y): ', 'y').lower()
    run_headless = headless != 'n'

    # ── File naming ───────────────────────────────────────────────────────
    safe_query    = re.sub(r'[^\w]', '_', query)[:30].strip('_')
    timestamp     = time.strftime('%Y%m%d_%H%M%S')
    base          = f'leads_{safe_query}_{timestamp}'

    print(f'\n  Query      : {query}')
    print(f'  Target     : {max_leads} leads')
    print(f'  Websites   : {"Yes" if do_web else "No"}')
    print(f'  Format     : {fmt.upper()}')
    print(f'  Output     : {base}.[xlsx/csv]')
    print(f'  Headless   : {"Yes" if run_headless else "No (browser window will open)"}')
    print('─' * 62)

    confirm = ask('\n✅ Start scraping? [y/n] (default y): ', 'y').lower()
    if confirm == 'n':
        print('Cancelled.')
        sys.exit(0)

    print()
    start_time = time.time()

    with sync_playwright() as pw:

        # ── Phase 1: Google Maps ──────────────────────────────────────────
        cprint(f'\n📍 PHASE 1 — Scraping Google Maps: "{query}"', '\033[1;34m')
        maps = GoogleMapsScraper(pw, headless=run_headless)
        businesses = maps.search(query, max_results=max_leads)

        if not businesses:
            print('❌ No results found. Tips:')
            print('   • Try a more specific query ("Coffee shops in Banjara Hills")')
            print('   • Re-run with headless=n to see what Google shows')
            print('   • Install playwright-stealth: pip install playwright-stealth')
            sys.exit(0)

        cprint(f'\n✅ Found {len(businesses)} businesses', '\033[1;32m')

        # ── Phase 2: Website scraping ─────────────────────────────────────
        if do_web:
            cprint(f'\n🌐 PHASE 2 — Scraping websites for contacts', '\033[1;34m')
            for i, biz in enumerate(businesses, 1):
                name    = (biz.get('title') or 'Unknown')[:38]
                website = (biz.get('website') or '').strip()

                if website and not any(d in website for d in SKIP_DOMAINS):
                    contacts = scrape_website_contacts(website, pw)
                    biz.update(contacts)

                    email = biz.get('email') or NIL
                    ig    = biz.get('instagram') or NIL
                    fb    = biz.get('facebook') or NIL
                    tags  = []
                    if email != NIL:
                        tags.append(f'📧 {email[:28]}')
                    if ig != NIL:
                        tags.append('📸 IG')
                    if fb != NIL:
                        tags.append('👍 FB')
                    detail = '  '.join(tags) or '✗ no contacts on site'
                else:
                    detail = '(no website to scrape)'

                print(f'  [{i:>3}/{len(businesses)}] {name:<38} {detail}')
                time.sleep(1.2)  # polite delay

        # ── Phase 3: Lead scoring ─────────────────────────────────────────
        cprint(f'\n📊 PHASE 3 — Scoring & ranking leads', '\033[1;34m')
        for biz in businesses:
            scores = score_lead(biz)
            biz.update(scores)

        businesses.sort(key=lambda x: x.get('lead_score', 0), reverse=True)

    # ── Phase 4: Save ─────────────────────────────────────────────────────
    elapsed = time.time() - start_time
    cprint(f'\n💾 PHASE 4 — Saving results  (ran for {elapsed:.0f}s)', '\033[1;34m')

    if fmt in ('excel', 'both') and EXCEL_AVAILABLE:
        save_excel(businesses, f'{base}.xlsx')
    elif fmt in ('excel', 'both') and not EXCEL_AVAILABLE:
        print('  ⚠️  Excel skipped (openpyxl not installed)')

    if fmt in ('csv', 'both'):
        save_csv(businesses, f'{base}.csv')

    # ── Summary ───────────────────────────────────────────────────────────
    print_summary(businesses)

    cprint(f'  Open  {base}.xlsx  to view & filter your leads!\n', '\033[1;32m')


if __name__ == '__main__':
    main()